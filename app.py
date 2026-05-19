"""
Torre de Control FTTH — Somos Internet  v2.0
Flask backend con roles, exportación, archivo mensual y KPIs históricos
"""

import os, re, json, csv, io, traceback, urllib.request, urllib.error
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from flask import (Flask, render_template, request, jsonify,
                   session, redirect, url_for, abort, Response, send_file)
from flask_sqlalchemy import SQLAlchemy

TZ_CO = ZoneInfo("America/Bogota")

def now_colombia():
    return datetime.now(TZ_CO).replace(tzinfo=None)

# ─── App setup ────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ftth-torre-control-2026-secreto')

raw_db = os.environ.get('DATABASE_URL', 'sqlite:///torre_control.db')
if raw_db.startswith('postgres://'):
    raw_db = raw_db.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = raw_db
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ─── Manejo de errores API ───────────────────────────

@app.errorhandler(Exception)
def handle_unexpected_error(e):
    """Devuelve errores reales en endpoints API para depurar desde el frontend."""
    if request.path.startswith('/api/') or request.path.startswith('/webhook/'):
        db.session.rollback()
        status = getattr(e, 'code', 500)
        return jsonify({
            'error': str(e),
            'type': e.__class__.__name__,
            'path': request.path,
            'trace': traceback.format_exc().splitlines()[-8:],
        }), status
    raise e


@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith('/api/') or request.path.startswith('/webhook/'):
        db.session.rollback()
        return jsonify({
            'error': str(e),
            'type': e.__class__.__name__,
            'path': request.path,
        }), 500
    return e

# ─── Modelos ──────────────────────────────────────────

class Tecnico(db.Model):
    __tablename__ = 'tecnicos'
    id        = db.Column(db.Integer, primary_key=True)
    nombre    = db.Column(db.String(100), nullable=False)
    zona      = db.Column(db.String(60))
    telefono  = db.Column(db.String(20))
    slack_id  = db.Column(db.String(20))
    activo    = db.Column(db.Boolean, default=True)
    creado    = db.Column(db.DateTime, default=now_colombia)
    tickets   = db.relationship('Ticket', backref='tecnico_rel', lazy=True)

    def to_dict(self):
        inicio_mes = now_colombia().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        tickets_mes = Ticket.query.filter(
            Ticket.tecnico_id == self.id,
            Ticket.fecha_apertura >= inicio_mes
        ).count()
        en_curso = Ticket.query.filter(
            Ticket.tecnico_id == self.id,
            Ticket.estado.in_(['EN_TRANSITO', 'EN_SITIO'])
        ).count()
        return {
            'id': self.id, 'nombre': self.nombre, 'zona': self.zona,
            'telefono': self.telefono, 'slack_id': self.slack_id,
            'activo': self.activo, 'tickets_mes': tickets_mes, 'en_curso': en_curso
        }


class Ticket(db.Model):
    __tablename__ = 'tickets'
    id              = db.Column(db.Integer, primary_key=True)
    slack_num       = db.Column(db.String(20))
    site            = db.Column(db.String(120), nullable=False)
    torre           = db.Column(db.String(40))
    zona            = db.Column(db.String(50))
    acc_nombre      = db.Column(db.String(120))
    acc_ip          = db.Column(db.String(20))
    edge_nombre     = db.Column(db.String(120))
    edge_ip         = db.Column(db.String(20))
    modelo          = db.Column(db.String(20))
    tipo            = db.Column(db.String(80))
    afectados       = db.Column(db.Integer, default=0)
    observacion     = db.Column(db.Text)
    macs            = db.Column(db.Text)
    topologia_url   = db.Column(db.Text)
    ubicacion_url   = db.Column(db.Text)
    appointment_num = db.Column(db.String(20))
    estado          = db.Column(db.String(20), default='ABIERTO')
    semaforo        = db.Column(db.String(10), default='VERDE')
    es_reincidente  = db.Column(db.Boolean, default=False)
    evento_num      = db.Column(db.Integer, default=1)
    tecnico_id      = db.Column(db.Integer, db.ForeignKey('tecnicos.id'), nullable=True)
    fecha_apertura  = db.Column(db.DateTime, default=now_colombia)
    fecha_asignacion= db.Column(db.DateTime)
    fecha_llegada   = db.Column(db.DateTime)
    fecha_cierre    = db.Column(db.DateTime)
    raw_mensaje     = db.Column(db.Text)
    cierre          = db.relationship('Cierre', backref='ticket', uselist=False, lazy=True,
                                      cascade='all, delete-orphan')

    @property
    def mttr_minutos(self):
        if self.fecha_apertura and self.fecha_cierre:
            return int((self.fecha_cierre - self.fecha_apertura).total_seconds() / 60)
        return None

    def to_dict(self):
        tec = self.tecnico_rel
        return {
            'id': self.id,
            'slack_num': self.slack_num,
            'site': self.site,
            'torre': self.torre,
            'zona': self.zona,
            'acc_nombre': self.acc_nombre,
            'acc_ip': self.acc_ip,
            'edge_nombre': self.edge_nombre,
            'edge_ip': self.edge_ip,
            'modelo': self.modelo,
            'tipo': self.tipo,
            'afectados': self.afectados,
            'observacion': self.observacion,
            'macs': json.loads(self.macs) if self.macs else [],
            'topologia_url': self.topologia_url,
            'ubicacion_url': self.ubicacion_url,
            'appointment_num': self.appointment_num,
            'estado': self.estado,
            'semaforo': self.semaforo,
            'es_reincidente': self.es_reincidente,
            'evento_num': self.evento_num,
            'tecnico_id': self.tecnico_id,
            'tecnico_nombre': tec.nombre if tec else None,
            'fecha_apertura': self.fecha_apertura.isoformat() if self.fecha_apertura else None,
            'fecha_asignacion': self.fecha_asignacion.isoformat() if self.fecha_asignacion else None,
            'fecha_llegada': self.fecha_llegada.isoformat() if self.fecha_llegada else None,
            'fecha_cierre': self.fecha_cierre.isoformat() if self.fecha_cierre else None,
            'mttr_minutos': self.mttr_minutos,
            'cierre': self.cierre.to_dict() if self.cierre else None,
        }


class Cierre(db.Model):
    __tablename__ = 'cierres'
    id              = db.Column(db.Integer, primary_key=True)
    ticket_id       = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    tecnico_id      = db.Column(db.Integer, db.ForeignKey('tecnicos.id'))
    causa_raiz      = db.Column(db.String(120))
    clasificacion   = db.Column(db.String(80))
    acciones        = db.Column(db.Text)
    estado_acc      = db.Column(db.String(40))
    estado_edge     = db.Column(db.String(40))
    estado_switch   = db.Column(db.String(40))
    potencia_dbm    = db.Column(db.Float)
    solucion        = db.Column(db.Text)
    estado_final    = db.Column(db.String(20))
    escalamiento    = db.Column(db.String(80))
    riesgo          = db.Column(db.Boolean, default=False)
    desc_riesgo     = db.Column(db.Text)
    recomendacion   = db.Column(db.Text)
    creado          = db.Column(db.DateTime, default=now_colombia)

    def to_dict(self):
        return {
            'causa_raiz': self.causa_raiz,
            'clasificacion': self.clasificacion,
            'acciones': json.loads(self.acciones) if self.acciones else [],
            'estado_acc': self.estado_acc,
            'estado_edge': self.estado_edge,
            'estado_switch': self.estado_switch,
            'potencia_dbm': self.potencia_dbm,
            'solucion': self.solucion,
            'estado_final': self.estado_final,
            'escalamiento': self.escalamiento,
            'recomendacion': self.recomendacion,
        }


class Actividad(db.Model):
    __tablename__ = 'actividad'
    id        = db.Column(db.Integer, primary_key=True)
    texto     = db.Column(db.Text)
    tipo      = db.Column(db.String(20))
    ticket_id = db.Column(db.Integer)
    usuario   = db.Column(db.String(80))
    ip        = db.Column(db.String(40))
    fecha     = db.Column(db.DateTime, default=now_colombia)

    def to_dict(self):
        return {
            'id': self.id, 'texto': self.texto, 'tipo': self.tipo,
            'ticket_id': self.ticket_id, 'usuario': self.usuario,
            'fecha': self.fecha.isoformat()
        }


class ArchivoMensual(db.Model):
    """Snapshots mensuales de datos — para historial permanente antes de limpieza."""
    __tablename__ = 'archivos_mensuales'
    id             = db.Column(db.Integer, primary_key=True)
    mes            = db.Column(db.String(7), unique=True, nullable=False)  # '2026-05'
    tickets_count  = db.Column(db.Integer, default=0)
    datos_json     = db.Column(db.Text)
    creado         = db.Column(db.DateTime, default=now_colombia)
    creado_por     = db.Column(db.String(80))

    def to_dict(self):
        return {
            'id': self.id, 'mes': self.mes,
            'tickets_count': self.tickets_count,
            'creado': self.creado.isoformat(),
            'creado_por': self.creado_por,
        }


# ─── Helpers ──────────────────────────────────────────

def log_actividad(texto, tipo, ticket_id=None, usuario=None):
    act = Actividad(
        texto=str(texto)[:300],
        tipo=tipo,
        ticket_id=ticket_id,
        usuario=usuario,
        ip=request.remote_addr if request else None
    )
    db.session.add(act)


def sync_ticket_to_sheets(ticket, action='upsert'):
    """Envía una copia del ticket a Google Sheets vía Apps Script.

    Es opcional: si GOOGLE_SHEETS_WEBHOOK_URL no existe, no hace nada.
    No debe romper la operación principal si Google falla.
    """
    url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '').strip()
    if not url:
        return

    try:
        payload = json.dumps({
            'action': action,
            'ticket': ticket.to_dict(),
            'synced_at': now_colombia().isoformat()
        }, ensure_ascii=False).encode('utf-8')

        req = urllib.request.Request(
            url,
            data=payload,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            if resp.status >= 400:
                print(f"[WARN] Google Sheets sync HTTP {resp.status}")
    except Exception as e:
        print(f"[WARN] No se pudo sincronizar ticket con Google Sheets: {e}")


def evaluar_reincidencia(ticket):
    inicio_mes = now_colombia().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    eventos_site = Ticket.query.filter(
        Ticket.site == ticket.site,
        Ticket.torre == ticket.torre,
        Ticket.fecha_apertura >= inicio_mes,
        Ticket.id != ticket.id
    ).count() + 1
    ticket.evento_num = eventos_site
    if eventos_site >= 3:
        ticket.semaforo = 'ROJO'; ticket.es_reincidente = True
    elif eventos_site >= 2:
        ticket.semaforo = 'AMARILLO'; ticket.es_reincidente = True
    else:
        eventos_acc = Ticket.query.filter(
            Ticket.acc_ip == ticket.acc_ip,
            Ticket.fecha_apertura >= inicio_mes,
            Ticket.id != ticket.id
        ).count()
        if eventos_acc >= 1:
            ticket.semaforo = 'AMARILLO'; ticket.es_reincidente = True
        else:
            ticket.semaforo = 'VERDE'; ticket.es_reincidente = False


def _periodo_inicio_fin(mes_str):
    """Retorna (inicio, fin) datetime para un mes '2026-05'."""
    inicio = datetime.strptime(mes_str, '%Y-%m')
    if inicio.month == 12:
        fin = datetime(inicio.year + 1, 1, 1)
    else:
        fin = datetime(inicio.year, inicio.month + 1, 1)
    return inicio, fin


def parse_slack_message(texto):
    """Parsea mensajes reales de Slack con formato de ticket FTTH."""
    data = {}
    texto = texto.replace('\xa0', ' ')
    texto = re.sub(r'\r\n?', '\n', texto)
    lines = [l.strip() for l in texto.split('\n') if l.strip()]

    collecting_obs = False
    obs_lines = []
    collecting_affected = False

    for i, line in enumerate(lines):
        clean = line.strip()
        low = clean.lower()

        if re.match(r'^(topolog[ií]a|ubicaci[oó]n|afectados:?|users:?|usuarios:?|t[eé]cnico:?|@)', clean, re.I):
            collecting_obs = False
        if collecting_obs:
            obs_lines.append(clean)
        if re.match(r'^(topolog[ií]a|ubicaci[oó]n|observaci[oó]n|tipo de problema|modelo equipo|nombre del sitio|torre\b|edge/acc|#\s*ticket|ticket\b)', clean, re.I):
            collecting_affected = False
        if re.match(r'^(afectados:?|users:?|usuarios:?)$', clean, re.I):
            collecting_affected = True; continue

        if collecting_affected and (re.match(r'^SOMOS-', clean, re.I) or re.search(r'[0-9a-f]{2}(:[0-9a-f]{2}){5}', clean, re.I)):
            data.setdefault('macs', [])
            mmac = re.search(r'([0-9a-f]{2}(?::[0-9a-f]{2}){5})', clean, re.I)
            if mmac:
                apt = re.search(r'APT\s*([A-Za-z0-9\-]+)', clean, re.I)
                data['macs'].append({'mac': mmac.group(1), 'apt': apt.group(1) if apt else None})
            else:
                data['macs'].append({'mac': clean, 'apt': None})

        m = re.search(r'Ticket\s*#?\s*(\d+)', clean, re.I)
        if m: data['slack_num'] = m.group(1)

        m = re.search(r'Edge/ACC\s+(COL_[A-Z0-9_]+)\s*-\s*([\d.]+)', clean, re.I)
        if m:
            nombre, ip = m.group(1).strip(), m.group(2).strip()
            if 'ACC' in nombre.upper(): data['acc_nombre'] = nombre; data['acc_ip'] = ip
            else: data['edge_nombre'] = nombre; data['edge_ip'] = ip
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                m2 = re.search(r'^(COL_[A-Z0-9_]+)\s*-\s*([\d.]+)', nxt, re.I)
                if m2:
                    n2, ip2 = m2.group(1).strip(), m2.group(2).strip()
                    if 'ACC' in n2.upper(): data['acc_nombre'] = n2; data['acc_ip'] = ip2
                    elif 'EDGE' in n2.upper(): data['edge_nombre'] = n2; data['edge_ip'] = ip2

        m = re.search(r'^(COL_[A-Z0-9_]+)\s*-\s*([\d.]+)', clean, re.I)
        if m:
            nombre, ip = m.group(1).strip(), m.group(2).strip()
            if 'ACC' in nombre.upper(): data['acc_nombre'] = nombre; data['acc_ip'] = ip
            elif 'EDGE' in nombre.upper(): data['edge_nombre'] = nombre; data['edge_ip'] = ip

        m = re.search(r'Nombre\s+del\s+sitio\s*:?-?\s*(.+)', clean, re.I)
        if m: data['site'] = m.group(1).strip()

        # Torre
        m = re.search(r'\bTorre\s*:?-?\s*([A-Za-z0-9\-]+)', clean, re.I)
        if m:
            data['torre'] = f"Torre {m.group(1).strip()}"

        # Zona
        m = re.search(r'\bZona\s*:?-?\s*(.+)', clean, re.I)
        if m:
            data['zona'] = m.group(1).strip()

        # Modelo equipo
        m = re.search(r'Modelo\s+equipo\s*:?-?\s*(.+)', clean, re.I)
        if m:
            data['modelo'] = m.group(1).strip()

        m = re.search(r'#\s*Afectados\s*:?-?\s*(\d+)', clean, re.I)
        if m: data['afectados'] = int(m.group(1))

        m = re.search(r'Tipo\s+de\s+Problema\s*:?-?\s*(.+)', clean, re.I)
        if m: data['tipo'] = m.group(1).strip()

        if re.match(r'^Observaci[oó]n\s*:?-?\s*$', clean, re.I):
            collecting_obs = True; continue
        m = re.search(r'Observaci[oó]n\s*:?-?\s*(.+)', clean, re.I)
        if m: data['observacion'] = m.group(1).strip()

        m = re.search(r'Appointment\s*#?\s*(\d+)', clean, re.I)
        if m: data['appointment_num'] = m.group(1)

        m = re.search(r'(https?://\S+)', clean)
        if m:
            url = m.group(1).strip()
            prev = lines[i - 1].lower() if i > 0 else ''
            ctx = low + ' ' + prev
            if 'drive' in url.lower() or 'topolog' in ctx: data['topologia_url'] = url
            elif 'maps' in url.lower() or 'goo.gl' in url.lower() or 'ubicac' in ctx: data['ubicacion_url'] = url

    if obs_lines and 'observacion' not in data:
        data['observacion'] = '\n'.join(obs_lines).strip()
    if 'afectados' not in data and data.get('macs'):
        data['afectados'] = len(data['macs'])
    return data


def require_login(roles=None):
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'rol' not in session:
                return redirect(url_for('login'))
            if roles and session['rol'] not in roles:
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return decorator


def _filtrar_tickets(desde=None, hasta=None):
    """Helper para filtrar tickets por rango de fechas."""
    q = Ticket.query
    if desde:
        q = q.filter(Ticket.fecha_apertura >= datetime.fromisoformat(desde))
    if hasta:
        q = q.filter(Ticket.fecha_apertura < datetime.fromisoformat(hasta) + timedelta(days=1))
    return q


# ─── Autenticación ────────────────────────────────────

@app.route('/')
def index():
    if 'rol' in session:
        return redirect(url_for('coordinador') if session['rol'] == 'coordinador' else url_for('tecnico'))
    return redirect(url_for('login'))


@app.route('/login')
def login():
    return render_template('login.html', tecnicos=Tecnico.query.filter_by(activo=True).order_by(Tecnico.nombre).all())


@app.route('/login', methods=['POST'])
def do_login():
    rol = request.form.get('rol')
    if rol == 'coordinador':
        if request.form.get('pin', '') != os.environ.get('COORDINADOR_PIN', '1234'):
            return render_template('login.html',
                                   tecnicos=Tecnico.query.filter_by(activo=True).all(),
                                   error='PIN incorrecto')
        session.update({'rol': 'coordinador', 'nombre': 'Coordinador', 'tecnico_id': None})
        return redirect(url_for('coordinador'))
    if rol == 'tecnico':
        tec = Tecnico.query.get(request.form.get('tecnico_id'))
        if not tec or not tec.activo:
            return render_template('login.html',
                                   tecnicos=Tecnico.query.filter_by(activo=True).all(),
                                   error='Técnico no encontrado')
        session.update({'rol': 'tecnico', 'tecnico_id': tec.id, 'nombre': tec.nombre})
        return redirect(url_for('tecnico'))
    return redirect(url_for('login'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── Vistas ───────────────────────────────────────────

@app.route('/coordinador')
@require_login(['coordinador'])
def coordinador():
    return render_template('coordinador.html', nombre=session['nombre'])


@app.route('/tecnico')
@require_login(['tecnico'])
def tecnico():
    return render_template('tecnico.html', tecnico=Tecnico.query.get(session['tecnico_id']), nombre=session['nombre'])


# ─── API: Parse Slack ─────────────────────────────────

@app.route('/api/parse-slack', methods=['POST'])
@require_login(['coordinador'])
def parse_slack_preview():
    raw = (request.json or {}).get('raw_mensaje', '').strip()
    if not raw: return jsonify({'error': 'Mensaje vacío'}), 400
    parsed = parse_slack_message(raw)
    if not parsed.get('site'): return jsonify({'error': 'No se pudo extraer el sitio', 'parsed': parsed}), 400
    return jsonify(parsed)


# ─── API: Tickets ─────────────────────────────────────

@app.route('/api/tickets', methods=['GET'])
@require_login()
def get_tickets():
    q = Ticket.query
    if session['rol'] == 'tecnico':
        q = q.filter(Ticket.tecnico_id == session['tecnico_id'])
    elif request.args.get('tecnico_id'):
        q = q.filter(Ticket.tecnico_id == int(request.args['tecnico_id']))
    if request.args.get('desde'):
        q = q.filter(Ticket.fecha_apertura >= datetime.fromisoformat(request.args['desde']))
    if request.args.get('hasta'):
        q = q.filter(Ticket.fecha_apertura < datetime.fromisoformat(request.args['hasta']) + timedelta(days=1))
    if request.args.get('estado') and request.args['estado'] != 'all':
        q = q.filter(Ticket.estado == request.args['estado'])
    return jsonify([t.to_dict() for t in q.order_by(Ticket.fecha_apertura.desc()).all()])


@app.route('/api/tickets', methods=['POST'])
@require_login(['coordinador'])
def crear_ticket():
    data = request.json or {}
    raw = data.get('raw_mensaje', '')
    if raw:
        parsed = parse_slack_message(raw)
        data.update({k: v for k, v in parsed.items() if k not in data or not data[k]})
    if not data.get('site'):
        return jsonify({'error': 'El campo site es obligatorio'}), 400
    slack_num = data.get('slack_num')
    if slack_num:
        ex = Ticket.query.filter_by(slack_num=str(slack_num)).first()
        if ex: return jsonify({'error': f'El ticket #{slack_num} ya existe', 'ticket': ex.to_dict()}), 409
    t = Ticket(
        slack_num       = slack_num,
        site            = data['site'],
        torre           = data.get('torre', 'Torre 1'),
        zona            = data.get('zona'),
        acc_nombre      = data.get('acc_nombre'),
        acc_ip          = data.get('acc_ip'),
        edge_nombre     = data.get('edge_nombre'),
        edge_ip         = data.get('edge_ip'),
        modelo          = data.get('modelo'),
        tipo            = data.get('tipo', 'Ausencia de Servicio'),
        afectados       = int(data.get('afectados', 0)),
        observacion     = data.get('observacion'),
        macs            = json.dumps(data.get('macs', [])),
        topologia_url   = data.get('topologia_url'),
        ubicacion_url   = data.get('ubicacion_url'),
        appointment_num = data.get('appointment_num'),
        raw_mensaje     = raw or None,
        fecha_apertura  = now_colombia(),
    )
    db.session.add(t); db.session.flush()
    evaluar_reincidencia(t)
    log_actividad(f"Nuevo ticket #{t.slack_num or t.id} — {t.site} {t.torre}", 'new', t.id, session.get('nombre'))
    db.session.commit()
    sync_ticket_to_sheets(t, 'created_manual')
    return jsonify(t.to_dict()), 201


@app.route('/api/tickets/<int:tid>', methods=['GET'])
@require_login()
def get_ticket(tid):
    t = Ticket.query.get_or_404(tid)
    if session['rol'] == 'tecnico' and t.tecnico_id != session['tecnico_id']: abort(403)
    return jsonify(t.to_dict())


@app.route('/api/tickets/<int:tid>/asignar', methods=['PUT'])
@require_login(['coordinador'])
def asignar_ticket(tid):
    t = Ticket.query.get_or_404(tid)
    tec = Tecnico.query.get_or_404((request.json or {}).get('tecnico_id'))
    t.tecnico_id = tec.id; t.estado = 'EN_TRANSITO'; t.fecha_asignacion = now_colombia()
    log_actividad(f"{tec.nombre} asignado a #{t.slack_num or t.id} — {t.site}", 'assign', t.id, session.get('nombre'))
    db.session.commit()
    sync_ticket_to_sheets(t, 'assigned')
    return jsonify(t.to_dict())


@app.route('/api/tickets/<int:tid>/en-sitio', methods=['PUT'])
@require_login()
def marcar_en_sitio(tid):
    t = Ticket.query.get_or_404(tid)
    if session['rol'] == 'tecnico' and t.tecnico_id != session['tecnico_id']: abort(403)
    t.estado = 'EN_SITIO'; t.fecha_llegada = now_colombia()
    log_actividad(f"Llegada a sitio — #{t.slack_num or t.id} {t.site}", 'assign', t.id, session.get('nombre'))
    db.session.commit()
    sync_ticket_to_sheets(t, 'en_sitio')
    return jsonify(t.to_dict())


@app.route('/api/tickets/<int:tid>/cerrar', methods=['POST'])
@require_login()
def cerrar_ticket(tid):
    t = Ticket.query.get_or_404(tid)
    if session['rol'] == 'tecnico' and t.tecnico_id != session['tecnico_id']: abort(403)
    data = request.json or {}
    missing = [f for f in ['causa_raiz', 'clasificacion', 'acciones', 'solucion', 'estado_final'] if not data.get(f)]
    if missing: return jsonify({'error': f'Faltan: {", ".join(missing)}'}), 400
    if len(data.get('solucion', '')) < 50: return jsonify({'error': 'Solución mínimo 50 caracteres'}), 400
    cierre = Cierre(
        ticket_id=t.id, tecnico_id=session.get('tecnico_id') or t.tecnico_id,
        causa_raiz=data['causa_raiz'], clasificacion=data['clasificacion'],
        acciones=json.dumps(data['acciones']), estado_acc=data.get('estado_acc'),
        estado_edge=data.get('estado_edge'), estado_switch=data.get('estado_switch'),
        potencia_dbm=data.get('potencia_dbm'), solucion=data['solucion'],
        estado_final=data['estado_final'], escalamiento=data.get('escalamiento'),
        riesgo=bool(data.get('riesgo')), desc_riesgo=data.get('desc_riesgo'),
        recomendacion=data.get('recomendacion'),
    )
    db.session.add(cierre)
    t.estado = data['estado_final']; t.fecha_cierre = now_colombia()
    if 'repetitiva' in data['clasificacion'].lower(): t.es_reincidente = True
    log_actividad(f"Ticket #{t.slack_num or t.id} {data['estado_final'].lower()} — {t.site} | {data['causa_raiz']}",
                  'closed' if data['estado_final'] == 'CERRADO' else 'alert', t.id, session.get('nombre'))
    db.session.commit()
    sync_ticket_to_sheets(t, 'closed')
    return jsonify(t.to_dict())


@app.route('/api/tickets/<int:tid>/estado', methods=['PUT'])
@require_login(['coordinador'])
def cambiar_estado_ticket(tid):
    t = Ticket.query.get_or_404(tid)
    data = request.json or {}
    nuevo_estado = data.get('estado')

    estados_validos = ['ABIERTO', 'EN_TRANSITO', 'EN_SITIO', 'ESCALADO']
    if nuevo_estado not in estados_validos:
        return jsonify({'error': 'Estado inválido. Usa ABIERTO, EN_TRANSITO, EN_SITIO o ESCALADO.'}), 400

    estado_anterior = t.estado
    t.estado = nuevo_estado

    if nuevo_estado == 'EN_TRANSITO' and not t.fecha_asignacion:
        t.fecha_asignacion = now_colombia()
    if nuevo_estado == 'EN_SITIO' and not t.fecha_llegada:
        t.fecha_llegada = now_colombia()
    if nuevo_estado == 'ABIERTO':
        t.fecha_llegada = None
        t.fecha_cierre = None

    log_actividad(
        f"Coordinador cambió estado de #{t.slack_num or t.id} de {estado_anterior} a {nuevo_estado} — {t.site}",
        'assign' if nuevo_estado in ['EN_TRANSITO', 'EN_SITIO'] else 'alert',
        t.id,
        session.get('nombre')
    )

    db.session.commit()
    sync_ticket_to_sheets(t, 'status_changed')
    return jsonify(t.to_dict())


@app.route('/api/tickets/<int:tid>', methods=['DELETE'])
@require_login(['coordinador'])
def eliminar_ticket(tid):
    """Solo coordinador puede eliminar tickets. Queda registro en auditoría."""
    t = Ticket.query.get_or_404(tid)
    ref = f"#{t.slack_num or t.id} — {t.site} {t.torre}"
    # cascade='all, delete-orphan' en el modelo borra el Cierre automáticamente
    db.session.delete(t)
    log_actividad(f"ELIMINAR ticket {ref}", 'alert', None, session.get('nombre'))
    db.session.commit()
    return jsonify({'ok': True, 'eliminado': ref})


# ─── API: Técnicos ────────────────────────────────────

@app.route('/api/tecnicos', methods=['GET'])
@require_login()
def get_tecnicos():
    return jsonify([t.to_dict() for t in Tecnico.query.order_by(Tecnico.nombre).all()])


@app.route('/api/tecnicos', methods=['POST'])
@require_login(['coordinador'])
def crear_tecnico():
    data = request.json or {}
    if not data.get('nombre'): return jsonify({'error': 'Nombre obligatorio'}), 400
    t = Tecnico(nombre=data['nombre'], zona=data.get('zona'), telefono=data.get('telefono'),
                slack_id=data.get('slack_id'), activo=True)
    db.session.add(t); db.session.commit()
    return jsonify(t.to_dict()), 201


@app.route('/api/tecnicos/<int:tid>', methods=['PUT'])
@require_login(['coordinador'])
def actualizar_tecnico(tid):
    t = Tecnico.query.get_or_404(tid)
    for f in ['nombre', 'zona', 'telefono', 'slack_id', 'activo']:
        if f in (request.json or {}): setattr(t, f, request.json[f])
    db.session.commit()
    return jsonify(t.to_dict())


@app.route('/api/tecnicos/<int:tid>', methods=['DELETE'])
@require_login(['coordinador'])
def desactivar_tecnico(tid):
    """Desactiva (soft delete) — el técnico ya no puede iniciar sesión."""
    t = Tecnico.query.get_or_404(tid)
    t.activo = False
    log_actividad(f"Técnico {t.nombre} desactivado", 'alert', None, session.get('nombre'))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/tecnicos/<int:tid>/eliminar', methods=['DELETE'])
@require_login(['coordinador'])
def eliminar_tecnico_permanente(tid):
    """Eliminación permanente — solo si no tiene tickets activos."""
    t = Tecnico.query.get_or_404(tid)
    activos = Ticket.query.filter(
        Ticket.tecnico_id == tid,
        Ticket.estado.in_(['ABIERTO', 'EN_TRANSITO', 'EN_SITIO'])
    ).count()
    if activos > 0:
        return jsonify({'error': f'El técnico tiene {activos} ticket(s) activo(s). Reasígnalos primero.'}), 400
    nombre = t.nombre
    db.session.delete(t)
    log_actividad(f"ELIMINAR técnico {nombre} (permanente)", 'alert', None, session.get('nombre'))
    db.session.commit()
    return jsonify({'ok': True})


# ─── API: KPIs generales ─────────────────────────────

@app.route('/api/kpis')
@require_login(['coordinador'])
def get_kpis():
    tickets = _filtrar_tickets(request.args.get('desde'), request.args.get('hasta')).all()
    total       = len(tickets)
    cerrados    = sum(1 for t in tickets if t.estado == 'CERRADO')
    reincid     = sum(1 for t in tickets if t.es_reincidente)
    sin_asignar = sum(1 for t in tickets if t.estado == 'ABIERTO' and not t.tecnico_id)
    afectados   = sum(t.afectados for t in tickets if t.estado != 'CERRADO')
    fcr = round(sum(1 for t in tickets if t.estado == 'CERRADO' and not t.es_reincidente) / cerrados * 100) if cerrados else 0
    mttrs = [t.mttr_minutos for t in tickets if t.mttr_minutos]
    mttr = round(sum(mttrs) / len(mttrs)) if mttrs else None

    site_counts = {}
    for t in tickets:
        k = f"{t.site} {t.torre}"
        if k not in site_counts: site_counts[k] = {'count': 0, 'semaforo': t.semaforo}
        site_counts[k]['count'] += 1
        if t.semaforo == 'ROJO': site_counts[k]['semaforo'] = 'ROJO'
        elif t.semaforo == 'AMARILLO' and site_counts[k]['semaforo'] != 'ROJO': site_counts[k]['semaforo'] = 'AMARILLO'

    falla_counts = {}
    for t in tickets:
        if t.tipo: falla_counts[t.tipo] = falla_counts.get(t.tipo, 0) + 1

    return jsonify({
        'total': total, 'abiertos': sum(1 for t in tickets if t.estado == 'ABIERTO'),
        'sin_asignar': sin_asignar, 'en_transito': sum(1 for t in tickets if t.estado == 'EN_TRANSITO'),
        'en_sitio': sum(1 for t in tickets if t.estado == 'EN_SITIO'), 'cerrados': cerrados,
        'afectados': afectados, 'reincidentes': reincid,
        'sites_rojo': len(set(f"{t.site} {t.torre}" for t in tickets if t.semaforo == 'ROJO')),
        'fcr': fcr, 'mttr': mttr,
        'tasa_reinc': round(reincid / total * 100) if total else 0,
        'sites': sorted(site_counts.items(), key=lambda x: -x[1]['count'])[:8],
        'fallas': sorted(falla_counts.items(), key=lambda x: -x[1])[:8],
    })


@app.route('/api/kpis/por-tecnico')
@require_login(['coordinador'])
def kpis_por_tecnico():
    """KPIs agrupados por técnico para el período seleccionado."""
    tickets = _filtrar_tickets(request.args.get('desde'), request.args.get('hasta')).all()
    tecnicos = {t.id: t for t in Tecnico.query.all()}
    stats = {}
    for tk in tickets:
        key = tk.tecnico_id or 0
        if key not in stats:
            tec = tecnicos.get(key)
            stats[key] = {'tecnico_id': key, 'nombre': tec.nombre if tec else 'Sin asignar',
                          'total': 0, 'cerrados': 0, 'reincidentes': 0, 'mttrs': [], 'afectados': 0}
        s = stats[key]
        s['total'] += 1
        if tk.estado == 'CERRADO': s['cerrados'] += 1
        if tk.es_reincidente: s['reincidentes'] += 1
        if tk.mttr_minutos: s['mttrs'].append(tk.mttr_minutos)
        s['afectados'] += tk.afectados

    result = []
    for s in stats.values():
        result.append({
            'tecnico_id': s['tecnico_id'], 'nombre': s['nombre'],
            'total': s['total'], 'cerrados': s['cerrados'],
            'fcr': round(s['cerrados'] / s['total'] * 100) if s['total'] else 0,
            'mttr': round(sum(s['mttrs']) / len(s['mttrs'])) if s['mttrs'] else None,
            'reincidentes': s['reincidentes'], 'afectados': s['afectados'],
        })
    return jsonify(sorted(result, key=lambda x: -x['total']))


@app.route('/api/kpis/tendencia')
@require_login(['coordinador'])
def kpis_tendencia():
    """Serie temporal de tickets: granularidad dia / semana / mes."""
    granularidad = request.args.get('granularidad', 'dia')
    tickets = _filtrar_tickets(request.args.get('desde'), request.args.get('hasta')).all()
    periods = {}
    for tk in tickets:
        if not tk.fecha_apertura: continue
        fa = tk.fecha_apertura
        if granularidad == 'mes':
            key = fa.strftime('%Y-%m')
        elif granularidad == 'semana':
            key = (fa - timedelta(days=fa.weekday())).strftime('%Y-%m-%d')
        else:
            key = fa.strftime('%Y-%m-%d')
        if key not in periods:
            periods[key] = {'periodo': key, 'total': 0, 'cerrados': 0, 'reincidentes': 0, 'afectados': 0}
        periods[key]['total'] += 1
        if tk.estado == 'CERRADO': periods[key]['cerrados'] += 1
        if tk.es_reincidente: periods[key]['reincidentes'] += 1
        periods[key]['afectados'] += tk.afectados
    return jsonify(sorted(periods.values(), key=lambda x: x['periodo']))


# ─── API: Actividad / Auditoría ───────────────────────

@app.route('/api/actividad')
@require_login()
def get_actividad():
    q = Actividad.query
    if request.args.get('desde'):
        q = q.filter(Actividad.fecha >= datetime.fromisoformat(request.args['desde']))
    if request.args.get('hasta'):
        q = q.filter(Actividad.fecha < datetime.fromisoformat(request.args['hasta']) + timedelta(days=1))
    acts = q.order_by(Actividad.fecha.desc()).limit(50).all()
    return jsonify([a.to_dict() for a in acts])


# ─── API: Exportación ─────────────────────────────────

def _tickets_para_export(desde, hasta):
    return _filtrar_tickets(desde, hasta).order_by(Ticket.fecha_apertura.desc()).all()


@app.route('/api/export/csv')
@require_login(['coordinador'])
def export_csv():
    """Exporta tickets a CSV (UTF-8 con BOM para Excel)."""
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    tickets = _tickets_para_export(desde, hasta)

    output = io.StringIO()
    output.write('﻿')  # BOM para Excel
    w = csv.writer(output)
    w.writerow(['ID', 'Ticket Slack', 'Fecha Apertura', 'Site', 'Torre',
                'ACC', 'IP ACC', 'Modelo', 'Tipo de Falla', 'Afectados',
                'Tecnico', 'Estado', 'Semaforo', 'Reincidente',
                'Fecha Asignacion', 'Fecha Llegada', 'Fecha Cierre', 'MTTR (min)',
                'Causa Raiz', 'Clasificacion', 'Solucion', 'Recomendacion'])

    for t in tickets:
        tec = t.tecnico_rel; c = t.cierre
        w.writerow([
            t.id, t.slack_num,
            t.fecha_apertura.strftime('%Y-%m-%d %H:%M') if t.fecha_apertura else '',
            t.site, t.torre, t.acc_nombre, t.acc_ip, t.modelo, t.tipo, t.afectados,
            tec.nombre if tec else '',
            t.estado, t.semaforo, 'Si' if t.es_reincidente else 'No',
            t.fecha_asignacion.strftime('%Y-%m-%d %H:%M') if t.fecha_asignacion else '',
            t.fecha_llegada.strftime('%Y-%m-%d %H:%M') if t.fecha_llegada else '',
            t.fecha_cierre.strftime('%Y-%m-%d %H:%M') if t.fecha_cierre else '',
            t.mttr_minutos or '',
            c.causa_raiz if c else '', c.clasificacion if c else '',
            c.solucion if c else '', c.recomendacion if c else '',
        ])

    fname = f"tickets_{desde or 'inicio'}_{hasta or 'hoy'}.csv"
    return Response(output.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/export/excel')
@require_login(['coordinador'])
def export_excel():
    """Exporta tickets + KPIs a Excel con formato."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    tickets = _tickets_para_export(desde, hasta)

    wb = Workbook()

    # ── Hoja 1: Tickets ──────────────────────────────────
    ws = wb.active
    ws.title = "Tickets"

    hdr_fill = PatternFill('solid', fgColor='1F6FEB')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    bdr = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = ['ID', 'Ticket Slack', 'Fecha Apertura', 'Site', 'Torre', 'ACC',
               'IP ACC', 'Modelo', 'Tipo de Falla', 'Afectados', 'Tecnico', 'Estado',
               'Semaforo', 'Reincidente', 'Fecha Cierre', 'MTTR (min)', 'Causa Raiz',
               'Clasificacion', 'Solucion']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    sem_colors = {'ROJO': 'FFDDDD', 'AMARILLO': 'FFFBCC', 'VERDE': 'DDFFDD'}

    for row_num, t in enumerate(tickets, 2):
        tec = t.tecnico_rel; c = t.cierre
        row = [
            t.id, t.slack_num,
            t.fecha_apertura.strftime('%Y-%m-%d %H:%M') if t.fecha_apertura else '',
            t.site, t.torre, t.acc_nombre, t.acc_ip, t.modelo, t.tipo, t.afectados,
            tec.nombre if tec else '',
            t.estado, t.semaforo, 'Si' if t.es_reincidente else 'No',
            t.fecha_cierre.strftime('%Y-%m-%d %H:%M') if t.fecha_cierre else '',
            t.mttr_minutos or '',
            c.causa_raiz if c else '', c.clasificacion if c else '',
            (c.solucion[:200] if c and c.solucion else ''),
        ]
        ws.append(row)
        fill_color = sem_colors.get(t.semaforo, 'FFFFFF')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill('solid', fgColor=fill_color)
            cell.border = bdr

    for col in ws.columns:
        max_w = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 45)
    ws.freeze_panes = 'A2'

    # ── Hoja 2: KPIs Resumen ─────────────────────────────
    ws2 = wb.create_sheet('KPIs Resumen')
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 20

    titulo_font = Font(bold=True, size=14, color='1F6FEB')
    ws2['A1'] = 'Reporte KPIs — Torre de Control FTTH'
    ws2['A1'].font = titulo_font
    ws2['A2'] = f'Periodo: {desde or "inicio"} al {hasta or "hoy"}'
    ws2['A2'].font = Font(italic=True, color='666666')

    total = len(tickets)
    cerrados = sum(1 for t in tickets if t.estado == 'CERRADO')
    reincid = sum(1 for t in tickets if t.es_reincidente)
    mttrs = [t.mttr_minutos for t in tickets if t.mttr_minutos]
    fcr = round(sum(1 for t in tickets if t.estado == 'CERRADO' and not t.es_reincidente) / cerrados * 100) if cerrados else 0

    kpis_rows = [
        ('Total tickets', total), ('Tickets cerrados', cerrados),
        ('FCR (1ra visita)', f'{fcr}%'),
        ('MTTR promedio', f'{round(sum(mttrs)/len(mttrs))} min' if mttrs else 'N/D'),
        ('Tasa reincidencia', f'{round(reincid/total*100)}%' if total else '0%'),
        ('Usuarios afectados', sum(t.afectados for t in tickets if t.estado != 'CERRADO')),
        ('Sites en ROJO', len(set(f"{t.site} {t.torre}" for t in tickets if t.semaforo == 'ROJO'))),
    ]

    for r, (label, val) in enumerate(kpis_rows, 4):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=val)

    # ── Hoja 3: Por Técnico ──────────────────────────────
    ws3 = wb.create_sheet('Por Tecnico')
    for col, h in enumerate(['Tecnico', 'Total', 'Cerrados', 'FCR%', 'MTTR (min)', 'Reincidencias'], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font

    tec_stats = {}
    for tk in tickets:
        key = tk.tecnico_id or 0
        tec = tk.tecnico_rel
        if key not in tec_stats:
            tec_stats[key] = {'nombre': tec.nombre if tec else 'Sin asignar', 'total': 0, 'cerrados': 0, 'reinc': 0, 'mttrs': []}
        tec_stats[key]['total'] += 1
        if tk.estado == 'CERRADO': tec_stats[key]['cerrados'] += 1
        if tk.es_reincidente: tec_stats[key]['reinc'] += 1
        if tk.mttr_minutos: tec_stats[key]['mttrs'].append(tk.mttr_minutos)

    for r, s in enumerate(sorted(tec_stats.values(), key=lambda x: -x['total']), 2):
        fcr_t = round(s['cerrados'] / s['total'] * 100) if s['total'] else 0
        mttr_t = round(sum(s['mttrs']) / len(s['mttrs'])) if s['mttrs'] else ''
        ws3.append([s['nombre'], s['total'], s['cerrados'], f'{fcr_t}%', mttr_t, s['reinc']])

    output = io.BytesIO(); wb.save(output); output.seek(0)
    fname = f"reporte_ftth_{desde or 'inicio'}_{hasta or 'hoy'}.xlsx"
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/export/pdf-report')
@require_login(['coordinador'])
def export_pdf_report():
    """Genera reporte HTML imprimible (usar Ctrl+P → Guardar como PDF en el navegador)."""
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    tickets = _tickets_para_export(desde, hasta)

    total = len(tickets)
    cerrados = sum(1 for t in tickets if t.estado == 'CERRADO')
    reincid = sum(1 for t in tickets if t.es_reincidente)
    mttrs = [t.mttr_minutos for t in tickets if t.mttr_minutos]
    mttr = round(sum(mttrs)/len(mttrs)) if mttrs else None
    fcr = round(sum(1 for t in tickets if t.estado == 'CERRADO' and not t.es_reincidente) / cerrados * 100) if cerrados else 0

    site_counts = {}
    for t in tickets:
        k = f"{t.site} / {t.torre}"; site_counts[k] = site_counts.get(k, 0) + 1
    top_sites = sorted(site_counts.items(), key=lambda x: -x[1])[:5]

    rows_html = ''
    for t in tickets[:100]:  # max 100 en el reporte
        tec = t.tecnico_rel
        sem_color = {'ROJO': '#FFDDDD', 'AMARILLO': '#FFFBCC', 'VERDE': '#DDFFDD'}.get(t.semaforo, '#FFF')
        rows_html += f"""<tr style="background:{sem_color}">
            <td>{t.slack_num or t.id}</td>
            <td>{t.fecha_apertura.strftime('%d/%m/%Y %H:%M') if t.fecha_apertura else ''}</td>
            <td>{t.site} / {t.torre}</td>
            <td>{t.tipo or ''}</td>
            <td style="text-align:center">{t.afectados}</td>
            <td>{tec.nombre if tec else '—'}</td>
            <td style="text-align:center">{t.estado.replace('_',' ')}</td>
            <td style="text-align:center">{t.mttr_minutos or '—'}</td>
        </tr>"""

    sites_rows = ''.join(f'<tr><td>{s}</td><td style="text-align:center">{c}</td></tr>' for s, c in top_sites)

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Reporte Torre de Control FTTH</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #222; padding: 20px; }}
  h1 {{ font-size: 18px; color: #1F6FEB; margin-bottom: 4px; }}
  h2 {{ font-size: 13px; color: #1F6FEB; margin: 16px 0 6px; border-bottom: 2px solid #1F6FEB; padding-bottom: 2px; }}
  .meta {{ color: #666; font-size: 10px; margin-bottom: 16px; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 10px; margin-bottom: 16px; }}
  .kpi {{ background: #EFF6FF; border: 1px solid #BFD7FF; border-radius: 6px; padding: 10px; text-align: center; }}
  .kpi .val {{ font-size: 24px; font-weight: 700; color: #1F6FEB; }}
  .kpi .lbl {{ font-size: 9px; color: #666; text-transform: uppercase; margin-top: 2px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
  th {{ background: #1F6FEB; color: #fff; padding: 5px 8px; text-align: left; }}
  td {{ padding: 4px 8px; border-bottom: 1px solid #EEE; }}
  .footer {{ text-align: center; color: #999; font-size: 9px; margin-top: 20px; }}
  @media print {{ button {{ display: none; }} }}
</style>
</head>
<body>
<button onclick="window.print()" style="background:#1F6FEB;color:#fff;border:none;padding:8px 16px;border-radius:4px;cursor:pointer;margin-bottom:16px">🖨 Imprimir / Guardar PDF</button>

<h1>🛰 Torre de Control FTTH — Reporte Operativo</h1>
<p class="meta">Período: {desde or 'Sin filtro'} al {hasta or 'Hoy'} &nbsp;|&nbsp; Generado: {now_colombia().strftime('%Y-%m-%d %H:%M')} &nbsp;|&nbsp; Por: {session.get('nombre','')}</p>

<h2>KPIs del Período</h2>
<div class="kpi-grid">
  <div class="kpi"><div class="val">{total}</div><div class="lbl">Total Tickets</div></div>
  <div class="kpi"><div class="val">{fcr}%</div><div class="lbl">FCR</div></div>
  <div class="kpi"><div class="val">{f'{mttr} min' if mttr else 'N/D'}</div><div class="lbl">MTTR Promedio</div></div>
  <div class="kpi"><div class="val">{round(reincid/total*100) if total else 0}%</div><div class="lbl">Tasa Reincidencia</div></div>
</div>

<h2>Top Sites más Reincidentes</h2>
<table><tr><th>Site / Torre</th><th>Tickets</th></tr>{sites_rows}</table>

<h2>Detalle de Tickets</h2>
<table>
<tr><th>Ticket</th><th>Fecha</th><th>Site/Torre</th><th>Tipo</th><th>Afect.</th><th>Técnico</th><th>Estado</th><th>MTTR</th></tr>
{rows_html}
</table>
{'<p style="color:#999;margin-top:8px">* Se muestran los primeros 100 tickets.</p>' if len(tickets) > 100 else ''}
<div class="footer">Torre de Control FTTH — Somos Internet &nbsp;|&nbsp; {now_colombia().strftime('%Y')}</div>
</body></html>"""

    return Response(html, mimetype='text/html')


# ─── API: Archivo Mensual ─────────────────────────────

@app.route('/api/historico/meses')
@require_login(['coordinador'])
def historico_meses():
    """Lista los meses que tienen datos de tickets."""
    from sqlalchemy import func, extract
    rows = db.session.query(
        extract('year', Ticket.fecha_apertura).label('anio'),
        extract('month', Ticket.fecha_apertura).label('mes_num'),
        func.count(Ticket.id).label('total')
    ).group_by('anio', 'mes_num').order_by('anio', 'mes_num').all()

    resultado = []
    for row in rows:
        mes_str = f"{int(row.anio):04d}-{int(row.mes_num):02d}"
        archivo = ArchivoMensual.query.filter_by(mes=mes_str).first()
        try:
            label = datetime.strptime(mes_str, '%Y-%m').strftime('%B %Y').capitalize()
        except Exception:
            label = mes_str
        resultado.append({
            'mes': mes_str, 'label': label, 'total': row.total,
            'archivado': bool(archivo),
            'fecha_archivo': archivo.creado.isoformat() if archivo else None,
            'creado_por': archivo.creado_por if archivo else None,
        })
    return jsonify(resultado)


@app.route('/api/historico/archivar', methods=['POST'])
@require_login(['coordinador'])
def archivar_mes():
    """Crea un snapshot JSON del mes. Debe hacerse ANTES de limpiar."""
    mes = (request.json or {}).get('mes')
    if not mes: return jsonify({'error': 'Mes requerido (formato: YYYY-MM)'}), 400

    inicio, fin = _periodo_inicio_fin(mes)
    tickets = Ticket.query.filter(Ticket.fecha_apertura >= inicio, Ticket.fecha_apertura < fin).all()

    tec_stats = {}
    for t in tickets:
        tec = t.tecnico_rel
        key = t.tecnico_id or 0
        if key not in tec_stats:
            tec_stats[key] = {'nombre': tec.nombre if tec else 'Sin asignar', 'total': 0, 'cerrados': 0}
        tec_stats[key]['total'] += 1
        if t.estado == 'CERRADO': tec_stats[key]['cerrados'] += 1

    mttrs = [t.mttr_minutos for t in tickets if t.mttr_minutos]
    datos = {
        'mes': mes, 'exportado_en': now_colombia().isoformat(),
        'exportado_por': session.get('nombre'),
        'total_tickets': len(tickets),
        'kpis': {
            'total': len(tickets),
            'cerrados': sum(1 for t in tickets if t.estado == 'CERRADO'),
            'reincidentes': sum(1 for t in tickets if t.es_reincidente),
            'afectados_total': sum(t.afectados for t in tickets),
            'mttr_promedio': round(sum(mttrs)/len(mttrs)) if mttrs else None,
        },
        'por_tecnico': list(tec_stats.values()),
        'tickets': [t.to_dict() for t in tickets],
    }

    archivo = ArchivoMensual.query.filter_by(mes=mes).first()
    if archivo:
        archivo.datos_json = json.dumps(datos, ensure_ascii=False)
        archivo.creado = now_colombia()
        archivo.creado_por = session.get('nombre')
        archivo.tickets_count = len(tickets)
    else:
        archivo = ArchivoMensual(mes=mes, tickets_count=len(tickets),
                                 datos_json=json.dumps(datos, ensure_ascii=False),
                                 creado_por=session.get('nombre'))
        db.session.add(archivo)

    log_actividad(f"Mes {mes} archivado — {len(tickets)} tickets", 'alert', None, session.get('nombre'))
    db.session.commit()
    return jsonify({'ok': True, 'mes': mes, 'tickets': len(tickets)})


@app.route('/api/historico/descargar/<mes>')
@require_login(['coordinador'])
def descargar_archivo(mes):
    """Descarga el snapshot JSON del mes archivado."""
    archivo = ArchivoMensual.query.filter_by(mes=mes).first_or_404()
    return Response(
        archivo.datos_json,
        mimetype='application/json; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=archivo_{mes}.json'}
    )


@app.route('/api/historico/limpiar', methods=['DELETE'])
@require_login(['coordinador'])
def limpiar_mes():
    """
    Elimina tickets de un mes SOLO si:
    1. El mes está archivado (snapshot generado)
    2. Se pasa confirmado=true explícitamente
    Nunca se ejecuta automáticamente.
    """
    data = request.json or {}
    mes = data.get('mes')
    if not mes: return jsonify({'error': 'Mes requerido'}), 400
    if not data.get('confirmado'): return jsonify({'error': 'Se requiere confirmado=true'}), 400

    archivo = ArchivoMensual.query.filter_by(mes=mes).first()
    if not archivo:
        return jsonify({'error': f'El mes {mes} NO está archivado. Archiva primero antes de limpiar.'}), 400

    inicio, fin = _periodo_inicio_fin(mes)
    tickets = Ticket.query.filter(Ticket.fecha_apertura >= inicio, Ticket.fecha_apertura < fin).all()
    count = len(tickets)
    for t in tickets:
        db.session.delete(t)

    log_actividad(f"LIMPIEZA: {count} tickets del mes {mes} eliminados", 'alert', None, session.get('nombre'))
    db.session.commit()
    return jsonify({'ok': True, 'eliminados': count, 'mes': mes})


# ─── Webhook Slack ────────────────────────────────────

@app.route('/webhook/slack', methods=['POST'])
def webhook_slack():
    expected = os.environ.get('WEBHOOK_TOKEN')
    if not expected: return jsonify({'error': 'WEBHOOK_TOKEN no configurado'}), 500
    if request.headers.get('X-Webhook-Token', '') != expected:
        return jsonify({'error': 'Token inválido'}), 401

    texto = (request.json or {}).get('text', '') or ''
    if not texto or 'Ticket' not in texto:
        return jsonify({'skipped': True}), 200

    parsed = parse_slack_message(texto)
    if not parsed.get('site'):
        return jsonify({'skipped': True, 'reason': 'No se pudo extraer site', 'parsed': parsed}), 200

    slack_num = parsed.get('slack_num')
    if slack_num:
        ex = Ticket.query.filter_by(slack_num=str(slack_num)).first()
        if ex: return jsonify({'created': False, 'duplicate': True, 'ticket_id': ex.id}), 200

    t = Ticket(
        slack_num       = slack_num,
        site            = parsed.get('site', 'Sin nombre'),
        torre           = parsed.get('torre', 'Torre 1'),
        zona            = parsed.get('zona'),
        acc_nombre      = parsed.get('acc_nombre'),
        acc_ip          = parsed.get('acc_ip'),
        edge_nombre     = parsed.get('edge_nombre'),
        edge_ip         = parsed.get('edge_ip'),
        modelo          = parsed.get('modelo'),
        tipo            = parsed.get('tipo', 'Ausencia de Servicio'),
        afectados       = int(parsed.get('afectados', 0)),
        observacion     = parsed.get('observacion'),
        macs            = json.dumps(parsed.get('macs', [])),
        topologia_url   = parsed.get('topologia_url'),
        ubicacion_url   = parsed.get('ubicacion_url'),
        appointment_num = parsed.get('appointment_num'),
        raw_mensaje     = texto,
        fecha_apertura  = now_colombia(),
    )
    
    db.session.add(t); db.session.flush()
    evaluar_reincidencia(t)
    log_actividad(f"Ticket #{t.slack_num or t.id} desde Slack — {t.site} {t.torre}", 'new', t.id, 'Slack Bot')
    db.session.commit()
    sync_ticket_to_sheets(t, 'created_slack')
    return jsonify({'created': True, 'ticket_id': t.id, 'slack_num': t.slack_num}), 201


# ─── Init ─────────────────────────────────────────────


def seed_db():
    pass  # DB inicia limpia. Agrega técnicos reales desde el panel Coordinador.


def ensure_schema_updates():
    """
    Actualizaciones pequeñas de esquema para PostgreSQL.
    db.create_all() crea tablas nuevas, pero no agrega columnas nuevas
    a tablas ya existentes.
    """
    try:
        with db.engine.begin() as conn:
            conn.exec_driver_sql(
                "ALTER TABLE tickets ADD COLUMN IF NOT EXISTS zona VARCHAR(50)"
            )
    except Exception as e:
        print(f"[WARN] No se pudo actualizar schema: {e}")


with app.app_context():
    db.create_all()
    ensure_schema_updates()
    seed_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
